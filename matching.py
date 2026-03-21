# matching.py
import re
import time
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import defaultdict

# ---------- Similarity threshold for partial match ----------
# Word-level Jaccard similarity must be >= this value to qualify as Partial.
# Keeps false positives like "Fund IV" vs "Fund I" or missing prefixes as No Match.
PARTIAL_SIMILARITY_THRESHOLD = 0.75

# Stopwords excluded from word overlap calculation
STOPWORDS = {"fund", "the", "of", "and", "a", "an", "for", "by", "in", "at"}

# ---------- Normalization helpers ----------
def collapse_spaces(s: str) -> str:
    return re.sub(r'\s+', ' ', s).strip()

def normalize_lp_cons(val):
    if val is None:
        return ""
    return collapse_spaces(str(val)).lower()

def normalize_fund(val):
    # Only collapses spaces and lowercases. Does NOT strip LP/LLC.
    # LP/LLC handling is done separately via strip_lp_llc and has_lp_llc.
    if val is None:
        return ""
    s = collapse_spaces(str(val))
    return s.lower()

def strip_lp_llc(s: str) -> str:
    """Remove LP/L.P./LLC/L.L.C from end of fund name, then clean trailing commas/punctuation."""
    s = re.sub(r'[\s,]*(l\.?p\.?|l\.?l\.?c\.?)[\s.,]*$', '', s, flags=re.IGNORECASE).strip()
    s = re.sub(r'[,.\s]+$', '', s).strip()
    return s

def has_lp_llc(s: str) -> bool:
    """Check if fund name contains LP or LLC (in any common format)."""
    return bool(re.search(r'\b(l\.?p\.?|l\.?l\.?c\.?)\b', s, flags=re.IGNORECASE))

def tokenize(s: str) -> list:
    """Split into lowercase alphanumeric tokens, excluding stopwords."""
    tokens = re.findall(r'[a-z0-9]+', s.lower())
    return [t for t in tokens if t not in STOPWORDS]

def word_jaccard(a: str, b: str) -> float:
    """
    Word-level Jaccard similarity between two strings.
    Jaccard = |intersection| / |union| of unique word sets.
    Returns 0.0 if either string is empty.
    """
    sa, sb = set(tokenize(a)), set(tokenize(b))
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / len(sa | sb)

def find_column_ignore_case(df, target_lower):
    for col in df.columns:
        if str(col).strip().lower() == target_lower:
            return col
    raise KeyError(f"Column '{target_lower}' not found.")

def read_uploaded(uploaded_bytes: bytes, filename: str):
    """Read uploaded file (CSV or Excel) into a DataFrame"""
    buffer = BytesIO(uploaded_bytes)
    if filename.lower().endswith(".csv"):
        try:
            return pd.read_csv(buffer, dtype=str, encoding="utf-8")
        except UnicodeDecodeError:
            buffer.seek(0)  # reset pointer
            return pd.read_csv(buffer, dtype=str, encoding="latin1")
    else:
        return pd.read_excel(buffer, dtype=str)

def process_files(master_bytes: bytes, output_bytes: bytes,
                  master_filename: str = "master.xlsx",
                  output_filename: str = "output.xlsx",
                  similarity_threshold: float = PARTIAL_SIMILARITY_THRESHOLD):
    """
    Run fund matching.
    Inputs: file bytes + filenames (to detect CSV vs Excel).
    Returns: (BytesIO result_xlsx, stats dict)

    Matching logic:
    - Exact:   stripped fund names match AND LP/LLC presence is the same on both sides
    - Partial: LP/LLC mismatch only (stripped names equal, but one has LP and other doesn't)
               OR word-level Jaccard similarity >= PARTIAL_SIMILARITY_THRESHOLD
    - No Match: everything else
    """

    # ---------- Read DataFrames ----------
    master_orig = read_uploaded(master_bytes, master_filename)
    output_orig = read_uploaded(output_bytes, output_filename)

    # ---------- Locate required columns ----------
    master_lp_col  = find_column_ignore_case(master_orig, "lp name")
    master_fund_col = find_column_ignore_case(master_orig, "fund name")
    master_cons_col = find_column_ignore_case(master_orig, "consultant")

    output_lp_col  = find_column_ignore_case(output_orig, "lpname")
    output_fund_col = find_column_ignore_case(output_orig, "fundname")
    output_cons_col = find_column_ignore_case(output_orig, "reportingconsultant")

    # ---------- Build normalized master structures ----------
    master_lp_norm   = master_orig[master_lp_col].fillna("").apply(normalize_lp_cons)
    master_fund_norm = master_orig[master_fund_col].fillna("").apply(normalize_fund)
    master_cons_norm = master_orig[master_cons_col].fillna("").apply(normalize_lp_cons)
    master_fund_orig = master_orig[master_fund_col].fillna("").astype(str)

    # LP/LLC-aware structures
    master_fund_raw      = master_orig[master_fund_col].fillna("").astype(str)
    master_fund_stripped = master_fund_raw.apply(lambda v: strip_lp_llc(normalize_fund(v)))
    master_fund_has_lp   = master_fund_raw.apply(lambda v: has_lp_llc(str(v)))

    exact_map        = {}
    funds_by_lp_cons = defaultdict(list)
    for lp_n, fund_n, fund_stripped, fund_has_lp, cons_n, fund_orig in zip(
        master_lp_norm, master_fund_norm, master_fund_stripped,
        master_fund_has_lp, master_cons_norm, master_fund_orig
    ):
        # Exact key: stripped fund + LP/LLC presence flag
        # "Blackrock L.P." and "Blackrock LP" -> same stripped name, both have LP -> Exact
        # "Blackrock LP" vs "Blackrock"        -> LP flag differs -> falls to Partial
        exact_map[(lp_n, fund_stripped, cons_n, fund_has_lp)] = fund_orig
        funds_by_lp_cons[(lp_n, cons_n)].append((fund_n, fund_stripped, fund_has_lp, fund_orig))

    # ---------- Normalize output ----------
    output_lp_norm   = output_orig[output_lp_col].fillna("").apply(normalize_lp_cons).tolist()
    output_fund_norm = output_orig[output_fund_col].fillna("").apply(normalize_fund).tolist()
    output_cons_norm = output_orig[output_cons_col].fillna("").apply(normalize_lp_cons).tolist()

    output_fund_raw      = output_orig[output_fund_col].fillna("").astype(str).tolist()
    output_fund_stripped = [strip_lp_llc(normalize_fund(v)) for v in output_fund_raw]
    output_fund_has_lp   = [has_lp_llc(v) for v in output_fund_raw]

    # ---------- Create workbook ----------
    tmp_buffer = BytesIO()
    output_orig.to_excel(tmp_buffer, index=False)
    tmp_buffer.seek(0)
    wb = load_workbook(tmp_buffer)
    ws = wb.active

    GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    masterentity_col = ws.max_column + 1
    flag_col         = masterentity_col + 1

    ws.cell(row=1, column=masterentity_col).value = "masterentity"
    ws.cell(row=1, column=flag_col).value         = "flag"

    exact_count    = 0
    partial_count  = 0
    no_match_count = 0
    log_lines      = []

    total = len(output_orig)
    t0    = time.time()

    for i in range(total):
        excel_row  = i + 2
        lp         = output_lp_norm[i]
        fund       = output_fund_norm[i]
        cons       = output_cons_norm[i]
        f_stripped = output_fund_stripped[i]
        f_has_lp   = output_fund_has_lp[i]

        matched_original_fund = None
        fill_color            = None
        flag_value            = None

        # --- Exact match ---
        if (lp, f_stripped, cons, f_has_lp) in exact_map:
            matched_original_fund = exact_map[(lp, f_stripped, cons, f_has_lp)]
            fill_color = GREEN
            flag_value = "Exact"
            exact_count += 1
            log_lines.append(f"Row {excel_row} | Exact -> {matched_original_fund}")

        else:
            # --- Partial match ---
            candidates = funds_by_lp_cons.get((lp, cons), [])
            for cand_norm, cand_stripped, cand_has_lp, cand_orig in candidates:

                # Case A: LP/LLC mismatch only — stripped names are identical, LP presence differs
                if f_stripped and cand_stripped and f_stripped == cand_stripped and f_has_lp != cand_has_lp:
                    matched_original_fund = cand_orig
                    fill_color = YELLOW
                    flag_value = "Partial"
                    partial_count += 1
                    log_lines.append(f"Row {excel_row} | Partial (LP mismatch) -> {matched_original_fund}")
                    break

                # Case B: Word-overlap similarity >= threshold (replaces loose substring match)
                # Uses stripped names so LP/LLC tokens don't inflate the score
                similarity = word_jaccard(f_stripped, cand_stripped)
                if similarity >= similarity_threshold:
                    matched_original_fund = cand_orig
                    fill_color = YELLOW
                    flag_value = "Partial"
                    partial_count += 1
                    log_lines.append(
                        f"Row {excel_row} | Partial (similarity={similarity:.2f}) -> {matched_original_fund}"
                    )
                    break

        # --- Write masterentity and flag columns ---
        if matched_original_fund:
            ws.cell(row=excel_row, column=masterentity_col).value = matched_original_fund
            ws.cell(row=excel_row, column=flag_col).value         = flag_value
        else:
            ws.cell(row=excel_row, column=masterentity_col).value = ""  # blank for No Match
            ws.cell(row=excel_row, column=flag_col).value         = "No Match"
            fill_color = RED
            no_match_count += 1
            log_lines.append(f"Row {excel_row} | No Match")

        # --- Apply row highlight ---
        if fill_color:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=excel_row, column=col).fill = fill_color

    elapsed = time.time() - t0

    # ---------- Save workbook into BytesIO ----------
    result_buffer = BytesIO()
    wb.save(result_buffer)
    result_buffer.seek(0)

    # ---------- Stats ----------
    stats = {
        "exact":   exact_count,
        "partial": partial_count,
        "nomatch": no_match_count,
        "rows":    total,
        "elapsed": elapsed,
        "log_lines": log_lines,
    }

    return result_buffer, stats
